from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


tmp = input("Path fichero de texto: ")
file = open(tmp)
tmp = str(input("Path excel: "))
wb = load_workbook(tmp)
ws = wb.active


def if_integer(textoA):
    try:
        int(textoA)
        return True
    except ValueError:
        return False


def toFloat(texto):
    numero = ""
    for i in range(len(texto)):
        if texto[i] == ",":
            numero += "."
        else:
            numero += texto[i]
    return float(numero)


if __name__ == "__main__":
    info = ""
    row = int(input("Fila inicio: "))
    col = int(input("Columna inicio: "))
    try:
        text = file.readline()
        while True:
            if not text:
                break
            if if_integer(text):
                ws[get_column_letter(col) + str(row)].value = int(text)
                row = row + 1
            text = file.readline()
    except Exception as e:
        print(e)
        file.close()
        wb.save(tmp)

    # End of while, close buffer and save info excel
    file.close()
    wb.save(tmp)
    print("Proceso finalizado con éxito :D")
    input("<<<<<<Pulse cualquier tecla para finalizar>>>>>>")