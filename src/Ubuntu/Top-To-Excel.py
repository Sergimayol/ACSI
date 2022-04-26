from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re

tmp = input("Path fichero de texto: ")
file = open(tmp)
tmp = str(input("Path excel: "))
wb = load_workbook(tmp)
ws = wb.active


def toFloat(texto):
    numero = ""
    for i in range(len(texto)):
        if texto[i] == ",":
            numero += "."
        else:
            numero += texto[i]
    return float(numero)


if __name__ == '__main__':
    row = 2
    col = 1
    while True:
        try:
            # CPu text
            text = file.read(8)
            ws[get_column_letter(col) + str(row)].value = text
            col = col + 1
            text = file.read(2)
            # User value
            text = file.read(3)
            ws[get_column_letter(col) + str(row)].value = toFloat(text)
            col = col + 1
            text = file.read(6)
            # System value
            text = file.read(3)
            ws[get_column_letter(col) + str(row)].value = toFloat(text)
            col = col + 1
            text = file.read(6)
            # Ni value
            text = file.read(3)
            ws[get_column_letter(col) + str(row)].value = toFloat(text)
            col = col + 1
            text = file.read(4)
            # id value
            text = file.read(5)
            ws[get_column_letter(col) + str(row)].value = toFloat(text)
            col = col + 1
            text = file.read(6)
            # wa value
            text = file.read(3)
            ws[get_column_letter(col) + str(row)].value = toFloat(text)
            col = col + 1
            text = file.read(6)
            # hi value
            text = file.read(3)
            ws[get_column_letter(col) + str(row)].value = toFloat(text)
            col = col + 1
            text = file.read(6)
            # si value
            text = file.read(3)
            ws[get_column_letter(col) + str(row)].value = toFloat(text)
            col = col + 1
            text = file.read(6)
            # st value
            text = file.read(3)
            ws[get_column_letter(col) + str(row)].value = toFloat(text)
            col = 1
            row = row + 1
            text = file.read(3)
            text = file.read(1)
            # Check EOF
            if not text:
                break
        except Exception as e:
            print(e)

    # End of while, close buffer and save info excel
    file.close()
    wb.save(tmp)
    print("Proceso finalizado con éxito :D")
    input("<<<<<<Pulse cualquier tecla para finalizar>>>>>>")
