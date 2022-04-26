from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re

tmp = input("Path fichero de texto: ")
file = open(tmp)
tmp = str(input("Path excel: "))
wb = load_workbook(tmp)
ws = wb.active

if __name__ == '__main__':
    text = file.read(1)
    info = ""
    row = 2
    col = 1
    while True:
        try:
            # Check EOF
            if not text:
                break
            # Add value
            if text == ":":
                ws[get_column_letter(col)+str(row)].value = info + ":"
                text = file.read(1)
                col = col + 1
                info = ""

            if text == ",":
                # Quitar a info los valores no númericos
                aux = [float(s) for s in re.findall(r'-?\d+\.?\d*', info)]
                ws[get_column_letter(col) + str(row)].value = aux[0]
                text = file.read(1)
                col = col + 1
                info = ""

            if text == "\n":
                # Quitar a info los valores no númericos
                aux = [float(s) for s in re.findall(r'-?\d+\.?\d*', info)]
                ws[get_column_letter(col) + str(row)].value = aux[0]
                # siguiente linea y reset columna e info
                row = row + 1
                col = 1
                info = ""

            # Add en info solo nums.
            info += text
            text = file.read(1)
        except Exception as e:
            print(e)

    # End of while, close buffer and save info excel
    file.close()
    wb.save(tmp)
    print("Proceso finalizado con éxito :D")
    input("<<<<<<Pulse cualquier tecla para finalizar>>>>>>")
